"""Finance-Report.xlsx — a generated multi-sheet VIEW for reading (not a master).

Requires openpyxl. If it is absent we degrade gracefully to a set of CSV views
under WORKDIR/report_csv/ and say so — the run still completes (§7).
"""

from __future__ import annotations

import csv
import os

from .model import TXN_COLUMNS


def _sheets(views, txns):
    """Return an ordered list of (sheet_name, header_row, data_rows)."""
    c = views["consolidated"]
    cf = views["cashflow"]
    out = []

    out.append(("Summary", ["metric", "value"], [
        ["Transactions", views["txn_count"]],
        ["Span from", views["generated_span"]["from"] or ""],
        ["Span to", views["generated_span"]["to"] or ""],
        ["Net position (excl. custodial)", c["net_position"]],
        ["Liquid total", c["liquid_total"]],
        ["Liabilities total", c["liability_total"]],
        ["Custodial (held in safekeeping)", c["custodial_total"]],
        ["Total in", cf["total_in"]],
        ["Total out", cf["total_out"]],
        ["Weekly burn", cf["weekly_burn"]],
        ["Daily burn", cf["daily_burn"]],
        ["Uncategorised spend", views["uncategorised_total"]],
    ]))
    out.append(("Consolidated",
                ["account_id", "name", "world", "type", "class", "balance", "custodial"],
                [[a["account_id"], a["name"], a["world"], a["type"], a["class"],
                  a["balance"], a["custodial"]] for a in c["accounts"]]))
    out.append(("Cashflow", ["month", "in", "out", "net"],
                [[m["month"], m["in"], m["out"], m["net"]] for m in cf["monthly"]]))
    out.append(("Categories", ["category", "amount"],
                [[r["category"], r["amount"]] for r in views["spend_by_category"]]))
    out.append(("Recurring", ["label", "cadence", "hits", "total", "expected"],
                [[r["label"], r["cadence"], r["hits"], r["total"], r["expected"]]
                 for r in views["recurring"]]))
    out.append(("Business-vs-Personal", ["scope", "in", "out"],
                [[r["scope"], r["in"], r["out"]] for r in views["business_personal"]["by_scope"]]))
    out.append(("Open-Questions", ["id", "type", "account_id", "detail", "answer"],
                [[q["id"], q["type"], q["account_id"], q["detail"], q["answer"]]
                 for q in views["open_questions"]]))
    out.append(("Transactions", TXN_COLUMNS,
                [[t.as_row()[c] for c in TXN_COLUMNS] for t in txns]))
    return out


def write_report(views, txns, workdir: str) -> tuple[str, str]:
    """Return (path, note). Falls back to CSV if openpyxl is unavailable."""
    sheets = _sheets(views, txns)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        out_dir = os.path.join(workdir, "report_csv")
        os.makedirs(out_dir, exist_ok=True)
        for name, header, rows in sheets:
            with open(os.path.join(out_dir, f"{name}.csv"), "w", newline="",
                      encoding="utf-8") as fh:
                w = csv.writer(fh)
                w.writerow(header)
                w.writerows(rows)
        return out_dir, "openpyxl not installed — wrote CSV views to report_csv/ instead"

    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    for name, header, rows in sheets:
        ws = wb.create_sheet(title=name[:31])
        ws.append(header)
        for cell in ws[1]:
            cell.font = bold
        for r in rows:
            ws.append(r)
    path = os.path.join(workdir, "Finance-Report.xlsx")
    wb.save(path)
    return path, ""
