"""Barclays current-account XLSX adapter (Personal + Joint in one workbook).

The workbook holds BOTH the personal and joint accounts in one table,
distinguished by an account-number column. We split them into two accounts by
looking each number up in accounts.csv (account_number -> account_id/world).
The real numbers live only in the local accounts.csv; a number we can't map is
flagged, never silently merged.

Requires openpyxl. If it's absent the adapter degrades loudly (a note the
console surfaces), it does not crash the run.
"""

from __future__ import annotations

import os

from ..model import RawTxn, IngestResult, money, parse_date
from .base import Adapter, register


def _norm(s) -> str:
    return str(s or "").strip().lower()


class BarclaysXlsxAdapter(Adapter):
    name = "barclays_xlsx"
    extensions = (".xlsx",)

    _AMOUNT_HINTS = ("amount", "value")
    _DATE_HINTS = ("date",)
    _DESC_HINTS = ("memo", "description", "narrative", "details", "reference")
    _ACCT_HINTS = ("account number", "account no", "account", "acc number")

    def _read_rows(self, path):
        """Read the whole sheet into memory and CLOSE the file (read_only mode on
        Windows otherwise holds the handle open). Returns (rows, error)."""
        try:
            import openpyxl  # noqa
        except ImportError:
            return None, "openpyxl not installed — cannot read Barclays XLSX (pip install openpyxl)"
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
        finally:
            wb.close()
        return rows, None

    def detect(self, path: str) -> bool:
        rows, err = self._read_rows(path)
        if not rows:
            return False
        headers = {_norm(c) for c in rows[0]}
        has_amount = any(any(h in c for h in self._AMOUNT_HINTS) for c in headers)
        has_acct = any(any(h in c for h in self._ACCT_HINTS) for c in headers)
        return has_amount and has_acct

    def _resolve_cols(self, headers: list[str]) -> dict:
        idx = {}
        for i, h in enumerate(headers):
            hn = _norm(h)
            if "amount" not in idx and any(x in hn for x in self._AMOUNT_HINTS):
                idx["amount"] = i
            if "date" not in idx and any(x in hn for x in self._DATE_HINTS):
                idx["date"] = i
            if "desc" not in idx and any(x in hn for x in self._DESC_HINTS):
                idx["desc"] = i
            # account: prefer the most specific hint that matches
            if any(x in hn for x in self._ACCT_HINTS):
                # keep the longest-matching header as the account column
                idx["account"] = i
        return idx

    def ingest(self, path, source_row, rules) -> IngestResult:
        base = os.path.basename(path)
        all_rows, err = self._read_rows(path)
        if err:
            return IngestResult([], [], [f"{base}: {err}"])
        if not all_rows:
            return IngestResult([], [], [f"{base}: empty workbook"])

        header = all_rows[0]
        rows = all_rows[1:]
        cols = self._resolve_cols([str(h) for h in header])
        if "amount" not in cols or "date" not in cols:
            return IngestResult([], [], [
                f"{base}: could not locate date/amount columns in {header}"])

        txns, notes = [], []
        unmapped_numbers = set()
        for n, row in enumerate(rows, start=2):
            if row is None or all(c is None for c in row):
                continue
            raw_date = row[cols["date"]] if cols["date"] < len(row) else None
            raw_amt = row[cols["amount"]] if cols["amount"] < len(row) else None
            if raw_date is None or raw_amt in (None, ""):
                continue
            try:
                iso = parse_date(raw_date)
                amount = money(raw_amt)
            except ValueError as e:
                notes.append(f"{base} row {n}: {e}; skipped")
                continue
            desc = ""
            if "desc" in cols and cols["desc"] < len(row):
                desc = str(row[cols["desc"]] or "").strip()
            desc = desc or "(no description)"

            # split into Personal vs Joint on the account-number column
            account_id, world = source_row.account_id, source_row.world
            if "account" in cols and cols["account"] < len(row):
                acct_num = str(row[cols["account"]] or "").strip()
                acc = rules.account_by_number.get(acct_num)
                if acc:
                    account_id, world = acc.account_id, acc.world
                elif acct_num:
                    unmapped_numbers.add(acct_num[-4:].rjust(len(acct_num), "*")
                                         if len(acct_num) > 4 else "****")
            txns.append(RawTxn(
                date=iso,
                amount=amount,
                description=desc,
                account_id=account_id,
                world=world,
                source_file=base,
                adapter=self.name,
            ))
        if unmapped_numbers:
            notes.append(
                f"{base}: {len(unmapped_numbers)} account number(s) not in accounts.csv "
                f"(masked: {', '.join(sorted(unmapped_numbers))}) — rows kept under "
                f"source default '{source_row.account_id}'; add the number->account map "
                f"to accounts.csv to split them.")
        return IngestResult(transactions=txns, control_totals=[], notes=notes)


register(BarclaysXlsxAdapter())
